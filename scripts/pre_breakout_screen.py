#!/usr/bin/env python3
"""
pre_breakout_screen.py — 起漲前整理型股票篩選腳本

用法：
  python3 pre_breakout_screen.py                         # 用最新交易日
  python3 pre_breakout_screen.py --date 20260610         # 指定交易日
  python3 pre_breakout_screen.py --save                  # 篩選後存檔

輸出：
  - 終端機印出：市場體檢 + 候選股列表
  - --save 時寫入：data/pre_breakout/pre_breakout_log-{MMDD}.json
  - --save 時同步匯出：data/pre_breakout/pre_breakout_log-{MMDD}.xlsx

資料依賴：
  - data/twse/2026/ 上市日K JSON（TWSE API 每日抓取）
  - data/tpex/2026/ 上櫃日K JSON（TPEX API 每日抓取）

篩選條件（保守模式，預設）：
  - 過去10天區間幅度 < 10%
  - 過去7天上漲天 ≤ 3
  - 當日漲跌幅（條件式）：
      * 多頭排列（MA5 > MA10）：閾值 7%，但 4%~7% 需檢查 40 日前高
      * 空頭/盤整（MA5 ≤ MA10）：閾值 4%（原規則）
  - 成交量 ≥ 1000 張
  - MA5 > MA10（多頭排列）
  - MA5 往上（今日MA5 > 前日MA5）
  - 收盤價 > MA5
  - 價格 > 10 元
  - 排除 ETF/權證（權值股/金融股不擋）

篩選條件（放寬模式，--relaxed）：
  '- 同上，但區間放寬到 < 25%、上漲天 ≤ 4
  - 權值股/高價股不擋（兩模式皆同）

分級（2026-06-17 更新，改以 dist_ma5 為主）：
  - A：dist_ma5 ≥ 3%（技術強勢，法人/量能支撐，首選）
  - B：1% ≤ dist_ma5 < 3%（技術面中等，次優先）
  - C：dist_ma5 < 1%（離 MA5 太近，觀察期）
"""

import json
import os
import sys
import glob
from copy import copy
from collections import OrderedDict
from datetime import datetime, timedelta

from portable_runtime import DATA_DIR, PROJECT_ROOT, load_dotenv

load_dotenv()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ModuleNotFoundError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None

# ── 路徑設定 ──
TWSE_DIR = os.path.join(str(DATA_DIR), 'twse', '2026')
TPEX_DIR = os.path.join(str(DATA_DIR), 'tpex', '2026')
PRE_BREAKOUT_DIR = os.path.join(str(DATA_DIR), 'pre_breakout')
OUTPUT_LATEST = os.path.join(PRE_BREAKOUT_DIR, 'pre_breakout_log.json')

if PatternFill is not None:
    HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
    HEADER_FONT = Font(color='FFFFFF', bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9E2F3'),
        right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'),
        bottom=Side(style='thin', color='D9E2F3'),
    )
    GRADE_FILLS = {
        'A': PatternFill(fill_type='solid', fgColor='E2F0D9'),
        'B': PatternFill(fill_type='solid', fgColor='FFF2CC'),
        'C': PatternFill(fill_type='solid', fgColor='FCE4D6'),
    }
else:
    HEADER_FILL = None
    HEADER_FONT = None
    THIN_BORDER = None
    GRADE_FILLS = {}

# ── 排除清單 ──
EXCLUDE_PREFIXES = ('00', '06', '07', '02', '03', '08', '91', '92', '93')
# 以下保留供參考，已不作用（2026-06-12 主人指示：不擋權值股/高價股）
HEAVY_WEIGHTS = {
    '2330', '2454', '2303', '2308', '2376', '2382', '2474',
    '3006', '3034', '3231', '3443', '3481', '3663', '4938',
    '5603', '6153', '6239', '6269', '6456', '6552', '6683',
    '6770', '8039', '8200', '8255', '8454', '8478',
    '8529', '8952', '8996', '9962', '9941',
}


def is_excluded(code, relaxed=False):
    """檢查股票代碼是否應排除（主人2026-06-12指示：不擋權值股/高價股）"""
    # 只排除 ETF、權證等非普通股
    if code.startswith(EXCLUDE_PREFIXES):
        return True
    return False


def normalize_field_name(name):
    text = str(name).replace(' ', '').strip()
    return text.replace('<br>', '').replace('<br/>', '').replace('<br />', '')


def is_valid_twse_file(path):
    try:
        with open(path, encoding='utf-8') as f:
            payload = json.load(f)
    except Exception:
        return False
    fields = [normalize_field_name(x) for x in payload.get('fields', [])]
    required = {'證券代號', '證券名稱', '成交股數', '收盤價'}
    return required.issubset(set(fields)) and bool(payload.get('data'))


def is_valid_tpex_file(path):
    try:
        with open(path, encoding='utf-8') as f:
            payload = json.load(f)
    except Exception:
        return False
    tables = payload.get('tables', [])
    if not tables:
        return False
    table = tables[0]
    fields = [normalize_field_name(x) for x in table.get('fields', [])]
    required = {'代號', '名稱', '成交股數', '收盤'}
    return required.issubset(set(fields)) and bool(table.get('data'))


def get_latest_date():
    """從有效共同交易日找出最新交易日。"""
    shared_dates = get_shared_dates()
    if not shared_dates:
        return None
    return shared_dates[-1]


def get_shared_dates():
    """取得 TWSE/TPEX 有效共同交易日（YYYYMMDD）。"""
    twse_dates = {
        os.path.basename(f)[:8]
        for f in glob.glob(os.path.join(TWSE_DIR, '*.json'))
        if is_valid_twse_file(f)
    }
    tpex_dates = {
        os.path.basename(f)[:8]
        for f in glob.glob(os.path.join(TPEX_DIR, '*.json'))
        if is_valid_tpex_file(f)
    }
    return sorted(twse_dates & tpex_dates)


def resolve_future_dates(target_date, lookahead=5):
    """取得 target_date 之後最多 lookahead 個共同交易日。"""
    dates = get_shared_dates()
    if target_date not in dates:
        return []
    idx = dates.index(target_date)
    return dates[idx + 1: idx + 1 + lookahead]


def load_market_snapshot(date_str, relaxed=False):
    """載入單一交易日快照，回傳 code -> {close, vol, market, name}。"""
    snapshot = {}

    twse_path = os.path.join(TWSE_DIR, f'{date_str}.json')
    if os.path.exists(twse_path):
        with open(twse_path, encoding='utf-8') as f:
            d = json.load(f)
        for row in d.get('data', []):
            try:
                code = str(row[0]).strip()
                if is_excluded(code, relaxed):
                    continue
                close = float(str(row[8]).replace(',', ''))
                if close <= 0:
                    continue
                if len(row) >= 16:
                    vol = int(str(row[2]).replace(',', '')) // 1000 if row[2] else 0
                else:
                    vol = int(str(row[1]).replace(',', '')) // 1000 if len(row) > 1 and row[1] else 0
                snapshot[code] = {
                    'date': date_str,
                    'close': close,
                    'vol': vol,
                    'name': str(row[1]).strip(),
                    'market': 'TWSE',
                }
            except (ValueError, IndexError, TypeError):
                continue

    tpex_path = os.path.join(TPEX_DIR, f'{date_str}.json')
    if os.path.exists(tpex_path):
        with open(tpex_path, encoding='utf-8') as f:
            d = json.load(f)
        for table in d.get('tables', []):
            for row in table.get('data', []):
                try:
                    code = str(row[0]).strip()
                    if is_excluded(code, relaxed):
                        continue
                    close = float(str(row[2]).replace(',', ''))
                    if close <= 0:
                        continue
                    vol = int(str(row[7]).replace(',', '')) // 1000 if row[7] else 0
                    snapshot[code] = {
                        'date': date_str,
                        'close': close,
                        'vol': vol,
                        'name': str(row[1]).strip(),
                        'market': 'TPEX',
                    }
                except (ValueError, IndexError, TypeError):
                    continue

    return snapshot


def build_future_days(code, base_close, target_date, relaxed=False):
    """建立往後最多 5 個交易日的收盤與雙漲幅資料。"""
    future_days = []
    prev_close = base_close
    for future_date in resolve_future_dates(target_date, lookahead=5):
        try:
            snapshot = load_market_snapshot(future_date, relaxed=relaxed)
        except Exception:
            continue
        row = snapshot.get(code)
        if not row:
            continue
        close = row['close']
        pct_from_signal = ((close - base_close) / base_close * 100) if base_close > 0 else None
        pct_from_prev = ((close - prev_close) / prev_close * 100) if prev_close > 0 else None
        future_days.append({
            'date': future_date,
            'close': round(close, 2),
            'pct_from_signal': round(pct_from_signal, 2) if pct_from_signal is not None else None,
            'pct_from_prev': round(pct_from_prev, 2) if pct_from_prev is not None else None,
        })
        prev_close = close
    return future_days


def load_data(target_date, relaxed=False):
    """載入指定交易日附近 13 個交易日的全市場資料"""
    result_file = os.path.join(TWSE_DIR, f'{target_date}.json')
    if not os.path.exists(result_file):
        print(f'找不到 TWSE 資料：{result_file}')
        return None

    # 找最近 13 個交易日檔案
    twse_files = sorted(glob.glob(os.path.join(TWSE_DIR, '*.json')))
    tpex_files = sorted(glob.glob(os.path.join(TPEX_DIR, '*.json')))

    # 定位 target_date 在檔案列表中的位置
    target_idx = None
    for i, f in enumerate(twse_files):
        if target_date in f:
            target_idx = i
            break
    if target_idx is None:
        print(f'找不到 {target_date} 在 TWSE 檔案列表中')
        return None

    # 取 target_date 前 13 個檔案（含 target_date 本身）
    start = max(0, target_idx - 12)
    recent_twse = twse_files[start:target_idx + 1]
    recent_tpex = [f for f in tpex_files
                   if any(d in f for d in [os.path.basename(x)[:8] for x in recent_twse])]
    recent_tpex = sorted(set(recent_tpex), key=lambda x: x)

    print(f'TWSE: {len(recent_twse)} 筆 ({os.path.basename(recent_twse[0])[:8]} ~ {target_date})')
    print(f'TPEX: {len(recent_tpex)} 筆')

    hist = OrderedDict()  # code -> [{name, market, close, vol, date, ...}]

    # TWSE 載入
    for fpath in recent_twse:
        fname = os.path.basename(fpath)[:8]
        with open(fpath, encoding='utf-8') as f:
            d = json.load(f)
        for row in d.get('data', []):
            try:
                code = str(row[0]).strip()
                if is_excluded(code, relaxed):
                    continue
                close = float(str(row[8]).replace(',', ''))
                if close <= 0:
                    continue
                # 成交量欄位：len=16 格式 index=2, len=10 格式 index=1
                if len(row) >= 16:
                    vol = int(str(row[2]).replace(',', '')) // 1000 if row[2] else 0
                else:
                    vol = int(str(row[1]).replace(',', '')) // 1000 if len(row) > 1 and row[1] else 0
                hist.setdefault(code, []).append({
                    'date': fname,
                    'open': float(row[5]),
                    'high': float(row[6]),
                    'low': float(row[7]),
                    'close': close,
                    'vol': vol,
                    'name': str(row[1]).strip(),
                    'market': 'TWSE',
                })
            except (ValueError, IndexError, TypeError):
                continue

    # TPEX 載入
    for fpath in recent_tpex:
        fname = os.path.basename(fpath)[:8]
        with open(fpath, encoding='utf-8') as f:
            d = json.load(f)
        for table in d.get('tables', []):
            for row in table.get('data', []):
                try:
                    code = str(row[0]).strip()
                    if is_excluded(code, relaxed):
                        continue
                    close = float(str(row[2]).replace(',', ''))
                    if close <= 0:
                        continue
                    vol = int(str(row[7]).replace(',', '')) // 1000 if row[7] else 0
                    hist.setdefault(code, []).append({
                        'date': fname,
                        'open': float(str(row[4]).replace(',', '')),
                        'high': float(str(row[5]).replace(',', '')),
                        'low': float(str(row[6]).replace(',', '')),
                        'close': close,
                        'vol': vol,
                        'name': str(row[1]).strip(),
                        'market': 'TPEX',
                    })
                except (ValueError, IndexError, TypeError):
                    continue

    print(f'載入 {len(hist)} 檔股票')
    return hist


def compute_40d_highs(target_date, hist_codes=None):
    """掃描 40 個交易日，回傳 {code: 40日內最高價}

    用於方案二前高檢查：pct 在 4%~7% 的股票需確認尚未突破前高。
    若 hist_codes 有傳入，只計算這些代碼（加速）。
    """
    twse_files = sorted(glob.glob(os.path.join(TWSE_DIR, '*.json')))
    tpex_files = sorted(glob.glob(os.path.join(TPEX_DIR, '*.json')))

    # 定位 target_date
    target_idx = None
    for i, f in enumerate(twse_files):
        if target_date in f:
            target_idx = i
            break
    if target_idx is None:
        return {}

    # 取 target_date 前 40 個檔案（不含 target_date 本身，因為是篩選日）
    start = max(0, target_idx - 40)
    recent_twse = twse_files[start:target_idx]  # 不包含 target_date
    twse_dates = {os.path.basename(f)[:8] for f in recent_twse}
    recent_tpex = sorted(set(
        f for f in tpex_files if os.path.basename(f)[:8] in twse_dates
    ))

    highs = {}

    for fpath in recent_twse:
        with open(fpath, encoding='utf-8') as f:
            d = json.load(f)
        for row in d.get('data', []):
            try:
                code = str(row[0]).strip()
                if is_excluded(code, False):
                    continue
                if hist_codes and code not in hist_codes:
                    continue
                high = float(row[6])
                if code not in highs or high > highs[code]:
                    highs[code] = high
            except (ValueError, IndexError, TypeError):
                continue

    for fpath in recent_tpex:
        with open(fpath, encoding='utf-8') as f:
            d = json.load(f)
        for table in d.get('tables', []):
            for row in table.get('data', []):
                try:
                    code = str(row[0]).strip()
                    if is_excluded(code, False):
                        continue
                    if hist_codes and code not in hist_codes:
                        continue
                    high = float(str(row[5]).replace(',', ''))
                    if code not in highs or high > highs[code]:
                        highs[code] = high
                except (ValueError, IndexError, TypeError):
                    continue

    return highs


def screen(hist, target_date, relaxed=False, highs_40d=None):
    """執行篩選，回傳候選股列表

    relaxed=True:
      - 放寬區間幅度到 25%（預設 10%）
      - 放寬上漲天數到 4（預設 3）
      - 含權值股/高價股（主人偏好：全部不擋）

    highs_40d: {code: 40日內最高價} — 用於方案二前高檢查

    方案二邏輯（2026-06-17 實作）：
      - 多頭排列（MA5 > MA10）時，pct 閾值放寬到 7%
      - 但 pct 在 4%~7% 的股票，需檢查 40 日前高（close < 前高）
      - 空頭/盤整（MA5 ≤ MA10）時，維持 4% 嚴格閾值
    """
    # 放寬參數
    RANGE_MAX = 25 if relaxed else 10
    UP_DAYS_MAX = 4 if relaxed else 3

    candidates = []

    def compute_rank_score(grade, dist_ma5, pct, vol_ratio, up_days):
        """排序分數：只改排序，不改篩選條件。"""
        grade_num = 2 if grade == 'A' else 1 if grade == 'B' else 0

        score = 0.0
        score += 2.0 * grade_num
        score += 0.6 * min(dist_ma5, 10)
        score += 0.2 * pct

        if dist_ma5 >= 6:
            score += 2.0
        elif dist_ma5 >= 4:
            score += 1.0

        if 2.5 <= pct <= 5.0:
            score += 1.2
        elif 5.0 < pct <= 7.0:
            score += 0.6

        if vol_ratio >= 1.2:
            score += 0.8

        if up_days == 3:
            score += 0.5
        elif up_days >= 4:
            score -= 0.5

        return round(score, 2)

    for code, data in hist.items():
        if len(data) < 12:
            continue

        latest = data[-1]
        name = latest['name']
        close = latest['close']
        vol = latest['vol']
        market = latest['market']

        # 價格門檻
        if close < 10:
            continue

        # 成交量 ≥ 1000 張（移到這裡做早期過濾）
        if vol < 1000:
            continue

        # 均線計算（需要在 pct 檢查之前，因為方案二需判斷 MA 多頭）
        closes = [d['close'] for d in data]
        vols = [d['vol'] for d in data]
        if len(closes) < 11:
            continue

        ma5_today = sum(closes[-6:-1]) / 5
        ma5_prev = sum(closes[-7:-2]) / 5
        ma10_today = sum(closes[-11:-1]) / 10

        # MA5 > MA10（多頭排列）
        if ma5_today <= ma10_today:
            continue
        # MA5 往上
        if ma5_today <= ma5_prev:
            continue
        # 收盤 > MA5
        if close <= ma5_today:
            continue

        # 漲跌幅計算
        prev_close = data[-2]['close']
        pct = (close - prev_close) / prev_close * 100 if prev_close > 0 else 0

        # === 方案二：條件式 pct 閾值 ===
        # 此時 ma5_today > ma10_today 已確認（多頭排列）
        if abs(pct) >= 7:
            # 多頭下閾值放寬到 7%，超過就排除
            continue
        if abs(pct) >= 4:
            # pct 4%~7%：需檢查 40 日前高
            high_40d = highs_40d.get(code) if highs_40d else None
            if high_40d is None:
                continue  # 查不到前高資料，保守排除
            if close >= high_40d:
                continue  # 已突破 40 日前高，排除
            # 距離前高還有空間 → 放行 ✅
        # pct < 4%：直接放行（原規則不變）

        dist_ma5 = (close - ma5_today) / ma5_today * 100

        # 過去10天區間幅度（不含今天）
        prices_10 = [d['close'] for d in data[-11:-1]]
        if prices_10:
            low_10 = min(prices_10)
            high_10 = max(prices_10)
            range_pct = (high_10 - low_10) / low_10 * 100 if low_10 > 0 else 0
        else:
            range_pct = 0
        if range_pct >= RANGE_MAX:
            continue

        # 過去7天上漲天數
        up_days = sum(
            1 for i in range(max(0, len(data) - 8), len(data) - 1)
            if data[i]['close'] > data[i - 1]['close']
        )
        if up_days > UP_DAYS_MAX:
            continue

        # 均量
        avg_vol_10 = sum(vols[-11:-1]) / 10 if len(vols) >= 11 else 0
        vol_ratio = vol / avg_vol_10 if avg_vol_10 > 0 else 0

        # 綜合分級（2026-06-17 主人確認）
        # A級：dist_ma5 ≥ 3%（技術強勢，離MA5有空間，法人/量能支撐）
        # B級：1% ≤ dist_ma5 < 3%（技術面中等，等待動能）
        # C級：dist_ma5 < 1%（離MA5太近，可能隨時跌破）
        if dist_ma5 >= 3:
            grade = 'A'
        elif dist_ma5 >= 1:
            grade = 'B'
        else:
            grade = 'C'

        # 主人 2026-06-27 指示：標準 / 保守選股都只保留 A 級
        if grade != 'A':
            continue

        rank_score = compute_rank_score(grade, dist_ma5, pct, vol_ratio, up_days)

        candidates.append({
            'code': code,
            'name': name,
            'market': market,
            'close': round(close, 2),
            'pct': round(pct, 2),
            'vol': vol,
            'ma5': round(ma5_today, 2),
            'ma10': round(ma10_today, 2),
            'dist_ma5': round(dist_ma5, 2),
            'range_pct': round(range_pct, 2),
            'up_days': up_days,
            'vol_ratio': round(vol_ratio, 2),
            'avg_vol_10': round(avg_vol_10),
            'institutional': {},
            'grade': grade,
            'rank_score': rank_score,
            'future_days': build_future_days(code, close, target_date, relaxed=relaxed),
        })

    # 排序（2026-06-27 更新）：只改排序，不改篩選條件
    # 優先依 rank_score 遞減；同分再看成交量
    candidates.sort(key=lambda c: (-c.get('rank_score', 0), -c['vol']))

    return candidates


def market_health_check(hist, target_date):
    """市場體檢：漲停家數、市場情緒"""
    limit_up = 0
    for code, data in hist.items():
        if not data:
            continue
        d = data[-1]
        prev = data[-2]['close'] if len(data) >= 2 else d['close']
        pct = (d['close'] - prev) / prev * 100 if prev > 0 else 0
        if 9.0 <= pct <= 10.5:
            limit_up += 1

    if limit_up > 40:
        mood = '熱絡'
        advice = '可正常執行'
    elif limit_up > 20:
        mood = '適中'
        advice = '可謹慎執行'
    else:
        mood = '冷淡'
        advice = '策略勝率僅 1.2%，建議暫不開倉'

    return {
        'limit_up_count': limit_up,
        'market_mood': mood,
        'advice': advice,
    }


def print_results(candidates, health, target_date, highs_40d=None):
    """印出結果到終端機"""
    print(f'\n{"=" * 60}')
    print(f'  PRE-BREAKOUT 篩選結果 — {target_date}')
    print(f'{"=" * 60}')
    print(f'  漲停家數：{health["limit_up_count"]} 檔 → {health["market_mood"]}')
    print(f'  建議：{health["advice"]}')
    print(f'  方案二：多頭 pct<=7% + 前高檢查（啟用）')
    print('  漲幅口徑：市場口徑=對前日（前收） | 研究口徑=對訊號日')
    print(f'{"=" * 60}')
    print(f'  通過篩選：{len(candidates)} 檔')
    print(f'{"=" * 60}')

    if not candidates:
        print('  (無符合條件的股票)')
        return

    for c in candidates:
        future_days = c.get('future_days') or []
        if future_days:
            future_text = ', '.join(
                f"{day['date']}:{day['close']:.2f}/{day['pct_from_signal']:+.2f}%/{day['pct_from_prev']:+.2f}%"
                for day in future_days
                if day.get('pct_from_signal') is not None and day.get('pct_from_prev') is not None
            ) or '(無後續資料)'
        else:
            future_text = '(無後續資料)'

        print(
            f"  {c['grade']} {c['code']} {c['name']} | "
            f"C={c['close']:.2f} V={c['vol']}張 分數={c.get('rank_score', 0):.2f} | 後5日={future_text}"
        )

    if health['limit_up_count'] <= 20:
        print(f'  注意：市場冷淡，以下僅供參考，不建議實際進場')

    # 分級統計
    counts = {'A': 0, 'B': 0, 'C': 0}
    for c in candidates:
        g = c.get('grade', 'C')
        if g in counts:
            counts[g] += 1
    print(f'  分級：A={counts["A"]}  B={counts["B"]}  C={counts["C"]}')


def get_next_trading_label(target_date):
    """回傳下一個交易日的 MM/DD 標籤；若資料檔尚未存在則用下一個平日推估。"""
    twse_files = sorted(glob.glob(os.path.join(TWSE_DIR, '*.json')))
    future_dates = [os.path.basename(f)[:8] for f in twse_files if os.path.basename(f)[:8] > target_date]
    if future_dates:
        dt = datetime.strptime(future_dates[0], '%Y%m%d')
        return f'{dt.month}/{dt.day}'

    dt = datetime.strptime(target_date, '%Y%m%d')
    while True:
        dt += timedelta(days=1)
        if dt.weekday() < 5:
            return f'{dt.month}/{dt.day}'


def style_header_row(ws, headers):
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx, header)
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = copy(THIN_BORDER)


def style_body_cell(cell, number_format=None):
    cell.border = copy(THIN_BORDER)
    if number_format:
        cell.number_format = number_format


def autosize_columns(ws, widths=None):
    if widths:
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        return
    for column in ws.columns:
        col_letter = column[0].column_letter
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def export_excel(output, dated_json_path):
    """把 pre_breakout log 同步匯出成固定格式 Excel。"""
    if Workbook is None:
        raise RuntimeError('缺少 openpyxl，無法匯出 Excel。請先安裝 requirements.txt，或改用不存檔模式執行。')

    target_date = output['date']
    today_dt = datetime.strptime(target_date, '%Y%m%d')
    today_label = f'{today_dt.month}/{today_dt.day}'
    next_label = get_next_trading_label(target_date)
    excel_path = dated_json_path.replace('.json', '.xlsx')

    candidates = output['candidates']
    rank = {'A': 0, 'B': 1, 'C': 2}
    sorted_rows = sorted(
        list(enumerate(candidates, start=1)),
        key=lambda x: (
            rank.get(x[1].get('grade', 'Z'), 99),
            -float(x[1].get('dist_ma5', 0) or 0),
            -float(x[1].get('vol', 0) or 0),
            x[1].get('code', ''),
        ),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = '候選股'

    candidate_headers = [
        '原始序號', '排序', '代號', '名稱', '市場', '分級',
        f'{today_label}收盤', f'{today_label}漲幅%',
        f'{next_label}收盤', f'{next_label}漲幅%',
        '成交量(張)', 'MA5', 'MA10', 'dist_MA5%', '區間%',
        '上漲天數', '量比', '10日均量', '投信買賣超',
        '自營買賣超', '外商買賣超', '法人合計',
    ]
    style_header_row(ws, candidate_headers)

    for sort_idx, (orig_idx, row) in enumerate(sorted_rows, start=1):
        inst = row.get('institutional') or {}
        trust = inst.get('investment_trust', 0) or 0
        dealer = inst.get('dealer', 0) or 0
        foreign = inst.get('foreign', 0) or 0
        values = [
            orig_idx, sort_idx, row.get('code'), row.get('name'), row.get('market'), row.get('grade'),
            row.get('close'), row.get('pct'), None, None,
            row.get('vol'), row.get('ma5'), row.get('ma10'), row.get('dist_ma5'), row.get('range_pct'),
            row.get('up_days'), row.get('vol_ratio'), row.get('avg_vol_10'), trust, dealer, foreign,
            trust + dealer + foreign,
        ]
        r = sort_idx + 1
        for c, value in enumerate(values, start=1):
            cell = ws.cell(r, c, value)
            number_format = None
            if c in (7, 8, 9, 10, 12, 13, 14, 15, 17):
                number_format = '0.00'
            elif c in (11, 16, 18, 19, 20, 21, 22):
                number_format = '#,##0.##'
            style_body_cell(cell, number_format=number_format)
        grade_fill = GRADE_FILLS.get(row.get('grade'))
        if grade_fill:
            ws.cell(r, 6).fill = copy(grade_fill)
            ws.cell(r, 6).alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:V{len(sorted_rows) + 1}'
    autosize_columns(ws, widths={
        'A': 10, 'B': 8, 'C': 10, 'D': 14, 'E': 10, 'F': 8,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 10,
        'M': 10, 'N': 12, 'O': 10, 'P': 10, 'Q': 10, 'R': 12,
        'S': 12, 'T': 12, 'U': 12, 'V': 12,
    })

    ws_sum = wb.create_sheet('摘要')
    style_header_row(ws_sum, ['項目', '數值'])
    grades = [r.get('grade') for r in candidates]
    dist_values = [r.get('dist_ma5') for r in candidates if r.get('dist_ma5') is not None]
    market_summary = output.get('market_summary') or {}
    summary_rows = [
        ('日期', output.get('date')),
        ('產出時間', output.get('generated')),
        ('總檔數', output.get('total', len(candidates))),
        ('A級', grades.count('A')),
        ('B級', grades.count('B')),
        ('C級', grades.count('C')),
        ('漲停家數', market_summary.get('limit_up_count')),
        ('市場情緒', market_summary.get('market_mood')),
        ('最高 dist_MA5', max(dist_values) if dist_values else None),
        ('最低 dist_MA5', min(dist_values) if dist_values else None),
        ('備註', f'今日檔案先填 {today_label}；{next_label} 收盤與漲幅欄位預留，待下個交易日收盤後補入'),
    ]
    for r, (label, value) in enumerate(summary_rows, start=2):
        ws_sum.cell(r, 1, label)
        ws_sum.cell(r, 2, value)
        style_body_cell(ws_sum.cell(r, 1))
        style_body_cell(ws_sum.cell(r, 2), number_format='0.00' if r in (10, 11) else None)
    ws_sum.freeze_panes = 'A2'
    autosize_columns(ws_sum, widths={'A': 18, 'B': 48})

    ws_desc = wb.create_sheet('欄位說明')
    style_header_row(ws_desc, ['欄位', '說明'])
    desc_rows = [
        ('原始序號', f'在 {target_date[4:]} JSON 中的原始順序'),
        ('排序', '為了分析方便重新排序（A/B/C + dist_MA5）'),
        (f'{today_label}收盤', '今日收盤價（來自 pre_breakout_log-MMDD.json）'),
        (f'{today_label}漲幅%', '今日 JSON 內的漲跌幅欄位'),
        (f'{next_label}收盤', '下個交易日收盤後再補入'),
        (f'{next_label}漲幅%', f'下個交易日收盤後用 {next_label} 與 {today_label} 收盤價計算'),
        ('dist_MA5%', '收盤價相對 MA5 的乖離率'),
        ('法人欄位', '單位：張（千股）；若 JSON 為空則先填 0'),
    ]
    for r, (label, value) in enumerate(desc_rows, start=2):
        ws_desc.cell(r, 1, label)
        ws_desc.cell(r, 2, value)
        style_body_cell(ws_desc.cell(r, 1))
        style_body_cell(ws_desc.cell(r, 2))
    ws_desc.freeze_panes = 'A2'
    autosize_columns(ws_desc, widths={'A': 18, 'B': 64})

    wb.save(excel_path)
    return excel_path


def save_results(candidates, health, target_date):
    """存檔到 pre_breakout 目錄"""
    os.makedirs(PRE_BREAKOUT_DIR, exist_ok=True)

    mmdd = target_date[4:]  # 0611

    # 檢查 dated 檔是否已存在
    dated_path = os.path.join(PRE_BREAKOUT_DIR, f'pre_breakout_log-{mmdd}.json')
    if os.path.exists(dated_path):
        print(f'  注意：{dated_path} 已存在，跳過存檔（防覆蓋）')
        return None

    output = {
        'date': target_date,
        'generated': target_date,
        'candidates': candidates,
        'total': len(candidates),
        'market_summary': {
            'limit_up_count': health['limit_up_count'],
            'market_mood': health['market_mood'],
            'tpex_pct_pass': True,
        },
    }

    # 寫 dated 檔
    with open(dated_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'  已存檔：{dated_path}')

    # 同時更新靜態檔（供 cron jobs 讀取）
    with open(OUTPUT_LATEST, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'  已更新：{OUTPUT_LATEST}')

    excel_path = export_excel(output, dated_path)
    print(f'  已匯出 Excel：{excel_path}')

    # 驗證
    assert len(output['candidates']) == output['total'], \
        f'total 與 candidates 長度不符：{output["total"]} vs {len(output["candidates"])}'
    assert os.path.exists(excel_path), f'Excel 匯出失敗：{excel_path}'

    return dated_path


def main():
    import argparse
    parser = argparse.ArgumentParser(description='起漲前整理型股票篩選')
    parser.add_argument('--date', help='交易日 (YYYYMMDD)，預設為最新')
    parser.add_argument('--save', action='store_true', help='儲存結果到 pre_breakout 目錄')
    parser.add_argument('--relaxed', action='store_true',
                        help='放寬模式：區間放寬到25%/上漲4天（權值股已全納入）')
    args = parser.parse_args()

    target_date = args.date or get_latest_date()
    if not target_date:
        print('找不到任何交易日資料')
        sys.exit(1)

    mode = '放寬' if args.relaxed else '保守'
    print(f'交易日：{target_date} | 模式：{mode} | 方案二：啟用')

    hist = load_data(target_date, relaxed=args.relaxed)
    if hist is None:
        sys.exit(1)

    # 計算 40 日前高（方案二用）
    print('計算 40 日前高...')
    highs_40d = compute_40d_highs(target_date)
    print(f'   完成：{len(highs_40d)} 檔股票有前高資料')

    candidates = screen(hist, target_date, relaxed=args.relaxed, highs_40d=highs_40d)
    health = market_health_check(hist, target_date)

    print_results(candidates, health, target_date, highs_40d=highs_40d)

    if args.save:
        saved = save_results(candidates, health, target_date)
        if saved:
            print(f'\n  執行：cat {saved} | python3 -m json.tool')

    print()


if __name__ == '__main__':
    main()
